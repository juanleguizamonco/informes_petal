import pandas as pd
import os
import openpyxl
from datetime import datetime
from copy import copy
from openpyxl.utils import get_column_letter
import requests

# Global variables for endpoints and storage paths
API_ENDPOINT = "http://localhost:8000/api/raw-data"  # Update with actual API endpoint
CATALOG_STORAGE = "catalogo_pay.xlsx"            # Update with actual storage path for catalog
TEMPLATE_STORAGE = "PayRegister_Template.xlsm"          # Update with actual storage path for output template

# --- Métodos para aplanar el JSON ---

def flatten_record(record: dict) -> dict:
    """
    Aplana un registro de JSON.
    Si existe la clave "paycomponents" y es una lista, extrae para cada componente
    su 'code', 'value' y 'ytdvalue', añadiendo nuevas claves al diccionario.
    Además, aplanará cualquier sub-diccionario encontrado en otras claves.
    """
    flat = {}
    for key, value in record.items():
        if key == "paycomponents" and isinstance(value, list):
            for comp in value:
                # Extrae el código; se asume que existe
                code = comp.get("code")
                if code is not None:
                    flat[code] = comp.get("value", None)
                    flat[f"YTD_{code}"] = comp.get("ytdvalue", None)
        elif isinstance(value, dict):
            # Aplanar sub-diccionario: concatenar clave principal y subclave
            for subkey, subval in value.items():
                flat[f"{key}__{subkey}"] = subval
        else:
            flat[key] = value
    return flat

def flatten_data(data: list) -> list:
    """
    Aplica flatten_record a cada elemento de la lista data.
    """
    return [flatten_record(item) for item in data]

# --- Carga y transformación de datos desde el API ---
try:
    response = requests.get(API_ENDPOINT)
    response.raise_for_status()
    raw_data = response.json()
    
    # Usar los métodos para aplanar el JSON
    flattened_data = flatten_data(raw_data)
    df_in = pd.DataFrame(flattened_data)
    print("Input DataFrame transformed and loaded successfully from API")
except Exception as e:
    print(f"Error loading and transforming input DataFrame from API: {e}")
    df_in = None

# --- Cargar catálogo desde Excel ---
try:
    df_catalogos = pd.read_excel(
        CATALOG_STORAGE, 
        sheet_name='Catalog', 
        engine='openpyxl')
    
    print("Catalog DataFrame loaded successfully")
except Exception as e:
    print(f"Error loading catalog DataFrame: {e}")
    df_catalogos = None

# --- Normalización y limpieza del DataFrame ---
df_flat = pd.json_normalize(df_in, sep='_')
df_flat = df_flat.applymap(lambda x: x.strip() if isinstance(x, str) else x)
df_clean = df_flat.replace({0: pd.NA, 0.0: pd.NA, "": pd.NA, "NaN": pd.NA})
df_input = df_clean.dropna(axis=1, how='all')
df_input = df_input.iloc[:, 6:]

# --- Cargar y preparar catálogo ---
df_catalogo_final = df_catalogos.iloc[:, 3:13]
df_catalogo_final.columns = df_catalogo_final.columns.str.strip()
df_catalogo_final = df_catalogo_final.dropna(how='all')

def obtener_periodo(df):
    """
    Extracts payroll period from the input DataFrame.
    """
    period_data = df['ME03210.000-ISGDDMFNN'].str.split(' ').str.get(1).dropna().unique()  # Ej: ME03210.000-ISGDDMFNN
    if len(period_data) == 1:
        return [str(period_data[0])]
    elif len(period_data) > 1:
        raise ValueError("Multiple periods found; expected only one.")
    return []

def obtener_catalogo_filtrado(catalog, df_input):
    """
    Filters catalog based on the payroll period from input data.
    """
    period = obtener_periodo(df_input)
    if not period:
        raise ValueError("No valid period found to filter the catalog.")
    return catalog[catalog['Uso'].isin(period)]

df_catalogo_filtrado = obtener_catalogo_filtrado(df_catalogo_final, df_input)

# ----- PAYREGISTER Processing -----
payregister_columns_static = {
    'Payroll Period': 'PERIOD/Payroll Type', # Añadir paycode
    'Entity': 'ME01410.000-ISGDDMFNN',
    'Pay group': 'ME05010.000-ISGDDMFNN',
    'Work Agreement': 'ME54710.000-ISGDDMFNN',
    'EE ID': 'ME01010.000-ISGDDMFNN',
    'Employee Name': 'ME02010.000-SSGDDMFNN'
}

mapping_pay = dict(zip(df_catalogo_filtrado['Paycode'], df_catalogo_filtrado['Encabezado']))
catalogo_order = df_catalogo_filtrado['Encabezado'].tolist()

df_input_static = df_input.rename(
    columns=lambda col: next((nuevo for nuevo, original in payregister_columns_static.items() if col == original), col)
)

df_input_renamed = df_input_static.rename(
    columns=lambda col: col if col in payregister_columns_static.keys() else mapping_pay.get(col, col)
)

nombres_validos_pay = set(payregister_columns_static.keys()).union(set(mapping_pay.values()))
df_filtrado = df_input_renamed[[col for col in df_input_renamed.columns if col in nombres_validos_pay]]

fixed_cols_order = [col for col in payregister_columns_static.keys() if col in df_filtrado.columns]
catalogo_cols_order = [col for col in catalogo_order if col in df_filtrado.columns and col not in fixed_cols_order]
extra_columns = df_catalogo_filtrado.sort_values(by='Sequence')['Encabezado'].tolist()
extra_columns = [col for col in extra_columns if col in df_filtrado.columns]

orden_final = []
for col in fixed_cols_order + extra_columns + catalogo_cols_order:
    if col not in orden_final:
        orden_final.append(col)

catalogo_register_mapping = dict(zip(df_catalogo_filtrado['Encabezado'], df_catalogo_filtrado['Register Type']))

df_output = df_filtrado[orden_final]
df_output = df_output.loc[:, ~df_output.columns.duplicated()]

register_type_row = {col: catalogo_register_mapping.get(col, '') for col in df_output.columns}

df_output_with_row = pd.concat(
    [pd.DataFrame([register_type_row], columns=df_output.columns), df_output],
    ignore_index=True
)

# ----- YTD REGISTER Processing -----
ytd_payregister_columns_static = {
    'Payroll Period': 'PERIOD/Payroll Type', #añadir el paycode
    'Entity': 'ME01410.000-ISGDDMFNN',
    'Pay group': 'ME05010.000-ISGDDMFNN',
    'Work Agreement': 'ME54710.000-ISGDDMFNN',
    'EE ID': 'ME01010.000-ISGDDMFNN',
    'Employee Name': 'ME02010.000-SSGDDMFNN'
}

mapping_ytd = dict(zip(df_catalogo_filtrado['YTD Paycode'], df_catalogo_filtrado['Encabezado']))

df_input_static_ytd = df_input.rename(
    columns=lambda col: next((nuevo for nuevo, original in ytd_payregister_columns_static.items() if col == original), col)
)

df_input_renamed_ytd = df_input_static_ytd.rename(
    columns=lambda col: col if col in ytd_payregister_columns_static.keys() else mapping_ytd.get(col, col)
)

nombres_validos_ytd = set(ytd_payregister_columns_static.keys()).union(set(mapping_ytd.values()))
df_filtrado_ytd = df_input_renamed_ytd[[col for col in df_input_renamed_ytd.columns if col in nombres_validos_ytd]]

fixed_cols_order_ytd = [col for col in ytd_payregister_columns_static.keys() if col in df_filtrado_ytd.columns]
catalogo_cols_order_ytd = [col for col in catalogo_order if col in df_filtrado_ytd.columns and col not in fixed_cols_order_ytd]
orden_final_ytd = fixed_cols_order_ytd + catalogo_cols_order_ytd

df_output_ytd = df_filtrado_ytd[orden_final_ytd]

# ----- Excel Formatting Functions -----
def copy_entire_sheet(source_sheet, target_sheet):
    """
    Copies all cells and their styles from source_sheet to target_sheet.
    """
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        if col_letter in target_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = col_dim.width
        else:
            target_sheet.column_dimensions[col_letter] = copy(col_dim)
    for row_index, row_dim in source_sheet.row_dimensions.items():
        if row_index in target_sheet.row_dimensions:
            target_sheet.row_dimensions[row_index].height = row_dim.height
        else:
            target_sheet.row_dimensions[row_index] = copy(row_dim)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.sheet_format = copy(source_sheet.sheet_format)

def data_format(df_output_with_row, df_output_ytd, folder_location, input_file_location, catalogs_file_location, template_file_location, country):
    """
    Formats data into the output Excel template.
    """
    ib = openpyxl.load_workbook(input_file_location, keep_vba=True)
    sheet_input_in = ib['Input']

    cb = openpyxl.load_workbook(catalogs_file_location, keep_vba=True)
    sheet_catalog = cb['Catalog']

    ob = openpyxl.load_workbook(template_file_location, keep_vba=True)
    sheet_input_out = ob['Input']
    sheet_catalog_out = ob['Catalog']
    sheet_payregister = ob['PayRegister']
    sheet_ytd_payregister = ob['YTDRegister']

    id_entidad = str(df_output_with_row['Entity'].iloc[1])
    dat = datetime.now().strftime('%Y%m%d')
    periodo = df_output_with_row['Payroll Period'].iloc[1].replace(' ', '_').replace('/', '_').replace('-', '_')

    ruta = os.path.join(os.getcwd(), folder_location)
    carpeta = os.path.join(ruta, f"{dat}_{country}_{id_entidad}_{periodo}_pre_Payregister")
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)
    s_filename = os.path.join(carpeta, f"{dat}_{country}_{id_entidad}_{periodo}_pre_Payregister.xlsm")

    copy_entire_sheet(sheet_input_in, sheet_input_out)
    copy_entire_sheet(sheet_catalog, sheet_catalog_out)

    register_type_row = list(df_output_with_row.iloc[0])
    num_cols = len(register_type_row)
    start_idx = 0
    for i in range(1, num_cols + 1):
        if i == num_cols or register_type_row[i] != register_type_row[start_idx]:
            group_value = register_type_row[start_idx]
            start_col_excel = start_idx + 2
            end_col_excel = i - 1 + 2
            if group_value not in [None, '']:
                if start_idx != i - 1:
                    col_start_letter = get_column_letter(start_col_excel)
                    col_end_letter = get_column_letter(end_col_excel)
                    merge_range = f"{col_start_letter}4:{col_end_letter}4"
                    sheet_payregister.merge_cells(merge_range)
                cell = sheet_payregister.cell(row=4, column=start_col_excel)
                cell.value = group_value
            start_idx = i

    for c_idx, col_name in enumerate(df_output_with_row.columns):
        cell = sheet_payregister.cell(row=5, column=c_idx + 2)
        cell.value = col_name

    for r_idx, row in df_output_with_row.iloc[1:].iterrows():
        for c_idx, value in enumerate(row):
            sheet_payregister.cell(row=r_idx + 5, column=c_idx + 2, value=value)

    for c_idx, col_name in enumerate(df_output_ytd.columns):
        cell = sheet_ytd_payregister.cell(row=5, column=c_idx + 2)
        cell.value = col_name

    for r_idx, row in df_output_ytd.iterrows():
        for c_idx, value in enumerate(row):
            sheet_ytd_payregister.cell(row=r_idx + 6, column=c_idx + 2, value=value)

    if os.path.exists(s_filename):
        try:
            os.remove(s_filename)
        except Exception as remove_error:
            print(f"Error removing existing file: {remove_error}")
    try:
        ob.save(s_filename)
    except Exception as e:
        print(f"Automatic save failed. Error: {e}")
        print(f"Please save the output file manually as: {s_filename}")

folder_location = "output_folder"
input_file_location = CATALOG_STORAGE
catalogs_file_location = CATALOG_STORAGE
template_file_location = TEMPLATE_STORAGE
country = 'MX'

data_format(df_output_with_row, df_output_ytd, folder_location, input_file_location, catalogs_file_location, template_file_location, country)


