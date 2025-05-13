import pandas as pd
import numpy as np
import os
import openpyxl
import requests
from datetime import datetime

# API endpoint for input JSON data
API_ENDPOINT = "http://localhost:8000/api/raw-data"  # Update with the actual API endpoint

# Storage path for catalogs Excel file
FILES_STORAGE = "catalogo_gl.xlsx"  # Update with the actual storage path

OUTPUT_STORAGE_GL = ""


# --- Load input data from API (JSON) and convert it to a flattened DataFrame ---

#Revisar si esta parte es mejor incluirla en el main del ETL
try:
    response = requests.get(API_ENDPOINT)
    response.raise_for_status()
    raw_data = response.json()

    # Aplanar los registros
    flattened_data = flatten_data(raw_data)
    df_in = pd.DataFrame(flattened_data)
    print("df_input loaded and flattened successfully from API")

except Exception as e:
    print(f"Error loading or flattening df_input from API: {e}")
    df_in = pd.DataFrame()  # evitar errores downstream


# --- Load catalogs data from Excel using the defined storage path ---
try:
    df_catalogos = pd.read_excel(
        FILES_STORAGE,
        sheet_name='Catalogs',
        engine='openpyxl',
        index_col=2
    )
    print("df_catalogos loaded successfully")
except Exception as e:
    print(f"Error loading df_catalogos: {e}")
    df_catalogos = pd.DataFrame()


# JSON flatted (normalización adicional y limpieza)
df_flat = pd.json_normalize(df_in, sep='_')
df_flat = df_flat.applymap(lambda x: x.strip() if isinstance(x, str) else x)
df_clean = df_flat.replace({0: pd.NA, 0.0: pd.NA, "": pd.NA, "NaN": pd.NA})
df_input = df_clean.dropna(axis=1, how='all').iloc[:, 6:]


# General catalogos format
df_catalogs = df_catalogos.copy().iloc[1:, :]
df_index = df_catalogs.iloc[1, :]
df_catalogs.columns = df_index
df_catalogs.reset_index(drop=True, inplace=True)

# Process catalog 1
df_catalogo1 = df_catalogs.iloc[2:, 2:14]

# Process catalog 2
df_catalogo2 = df_catalogs.iloc[2:, 16:20].dropna()

# Process catalog 3
df_catalogo3 = df_catalogs.iloc[2:, 22:29]

# Define detailed columns for the analysis DataFrame
detailed_columns = [
    "Payroll Period", "Entity", "Pay group", "Work Agreement", "EE ID", "Employee Name",
    "Client ID", "Client Description", "Cost Center", "% CC", "Level2 (Optional)",
    "Level3 (Optional)", "Level4 (Optional)", "Amount", "Debt Account", "Acred Account",
    "Item Text", "GL File", "Observations"
]
df_detailed = pd.DataFrame(columns=detailed_columns)


def crear_fila(period, entity, pay_group, work_agreement, employee_id, employee_name,
               client_id, client_description, cost_center, percentage_cc,
               level2, level3, level4, amount, debit_account, acred_account,
               item_text, gl_file):
    """Creates a dictionary representing a row for the analysis DataFrame."""
    return {
        'Payroll Period': period,
        "Entity": entity,
        "Pay group": pay_group,
        "Work Agreement": work_agreement,
        "EE ID": employee_id,
        "Employee Name": employee_name,
        "Client ID": client_id,
        "Client Description": client_description,
        "Cost Center": cost_center,
        "% CC": percentage_cc,
        "Level2 (Optional)": level2,
        "Level3 (Optional)": level3,
        "Level4 (Optional)": level4,
        "Amount": amount,
        "Debt Account": debit_account,
        "Acred Account": acred_account,
        "Item Text": item_text,
        "GL File": gl_file,
        "Observations": ""
    }

def obtener_entidad(df_prorrateo):
    """Returns a list of unique entities from the input DataFrame."""
    return df_prorrateo['AY01200.000-ISGDDMFNN'].dropna().unique().tolist()

def obtener_periodo(df_general):
    """Extracts and returns the payroll period from the input DataFrame."""
    period_data = df_general['ME03210.000-ISGDDMFNN'].str.split(' ').str[1].dropna().unique() #ME03210.000-ISGDDMFNN
    if len(period_data) == 1:
        return [str(period_data[0])]
    elif len(period_data) > 1:
        raise ValueError("Multiple periods found; expected only one.")
    return []

period_data = obtener_periodo(df_input)

def eliminar_nans(cadena):
    """Removes 'nan' strings from the input string."""
    if isinstance(cadena, str):
        return ', '.join(filter(lambda x: x.lower() != 'nan', cadena.split(', ')))
    return cadena

def obtener_centros_de_costos(employee_data, df_catalogo2):
    centro_de_costos_prorrateo = []
    centro_de_costos_solo = []
    
    for employee_id in employee_data['ME01010.000-ISGDDMFNN']: #ME01010.000-ISGDDMFNN
        if employee_id in df_catalogo2['ID'].values:
            porcentajes = df_catalogo2.loc[df_catalogo2['ID'] == employee_id, '% DE CC'].tolist()
            cc_values = df_catalogo2.loc[df_catalogo2['ID'] == employee_id, 'CC'].tolist()
            
            for porcentaje, cc_value in zip(porcentajes, cc_values):
                if porcentaje > 0:  # Solo añadir si el porcentaje es mayor que 0
                    centro_de_costos_prorrateo.append([employee_id, float(porcentaje) / 100, cc_value])
                else:
                    centro_de_costos_solo.append([employee_id, 1.0, None])  # Valor por defecto si no hay coincidencias
            
    return centro_de_costos_prorrateo, centro_de_costos_solo

def calcular_amount(valor, porcentaje, catalogo, paycode, employee_id, df_catalogo3,
                    period, entity, pay_group, work_agreement, employee_name, cost_center,
                    level2, level3, level4):
    """Calculates the amount based on the provided parameters."""
    paycode_str = str(paycode).strip()
  
    catalogo['Petal Code'] = catalogo['Petal Code'].astype(str).str.strip()
    filas_catalogo = catalogo[catalogo['Petal Code'] == paycode_str]
    
    if filas_catalogo.empty:
        return pd.DataFrame()
    
    base_amount = abs(float(valor) * porcentaje)
    amounts = []
    for _, fila in filas_catalogo.iterrows():
        accounting_type = fila.get('Accounting type')
        debit_account = fila.get('DEBIT account')
        credit_account = fila.get('CREDIT account')
        client_code = fila.get('Client Code')
        campo = fila.get('Campo')
        gl_file = fila.get('GLFile')
        item_text = fila.get('Item TEXT')
        
        if pd.notna(debit_account) or pd.notna(credit_account):
            if accounting_type in ['D-', 'C+']:
                amount = -base_amount
            elif accounting_type in ['P+', 'N+']:
                amount = base_amount
                df_temp = df_catalogo3[df_catalogo3['Employee ID'] == employee_id] #ME01010.000-ISGDDMFNN
                print(f"Catalog matches in df_catalogo3 for Employee ID {employee_id}: {len(df_temp)}")
                for _, row3 in df_temp.iterrows():
                    client_id = f"{row3['Employee ID']} {row3['Tipo']}" #ME01010.000-ISGDDMFNN
                    client_description = row3['BENEFICIARIA']
                    acred_account = row3['NUM. ACREEDOR']
                    amounts.append({
                        'Amount': amount,
                        'DEBIT account': '', 
                        'CREDIT account': acred_account,
                        'Client Code': client_id,
                        'Campo': client_description,
                        'GLFile': gl_file,
                        'Item Text': item_text
                    })
                continue
            else:
                amount = base_amount
            amounts.append({
                'Amount': amount,
                'DEBIT account': debit_account,
                'CREDIT account': credit_account,
                'Client Code': client_code,
                'Campo': campo,
                'GLFile': gl_file,
                'Item Text': item_text
            })
    return pd.DataFrame(amounts)

def procesar_empleado(employee_data, df_catalogo_filtrado, df_catalogo3):
    """Processes an employee's data to generate analysis rows."""
    analysis = []
    period = str(employee_data['ME03210.000-ISGDDMFNN'].iloc[0]) #ME03210.000-ISGDDMFNN
    entity = employee_data['AY01200.000-ISGDDMFNN'].iloc[0]
    pay_group = (str(employee_data['ME05010.000-ISGDDMFNN'].iloc[0])
                 if 'ME05010.000-ISGDDMFNN' in employee_data.columns and not employee_data["ME05010.000-ISGDDMFNN"].empty
                 else 'Paygroup not found')
    work_agreement = (str(employee_data['ME54710.000-ISGDDMFNN'].iloc[0])
                      if 'ME54710.000-ISGDDMFNN' in employee_data.columns and not employee_data['ME54710.000-ISGDDMFNN'].empty
                      else 'Work agreement not found')
    employee_name = (str(employee_data['ME02010.000-SSGDDMFNN'].iloc[0])
                     if 'ME02010.000-SSGDDMFNN' in employee_data.columns and not employee_data['ME02010.000-SSGDDMFNN'].empty
                     else "")
    employee_id = int(employee_data['ME01010.000-ISGDDMFNN'].values[0]) #ME01010.000-ISGDDMFNN
    
    centro_de_costos_prorrateo, centro_de_costos_solo = obtener_centros_de_costos(employee_data, df_catalogo2)

    processed_keys = set()
    
    level2 = level3 = level4 = ''

    base_columns = ['PERIOD/Payroll Type', 'Legal','EMPLOYEE_ID','CO30720.000-IFGDNACNN','CO30730.000-IFGDNACNN','ME01010.000-ISGDDMFNN','ME01410.000-ISGDDMFNN','ME02010.000-SSGDDMFNN','ME41210.000-ISGDDMFNN']

    for paycode in employee_data.columns:

        if paycode in base_columns:
            continue

        value = employee_data.iloc[0][paycode]
        if pd.isna(value):
            continue

        # Verificar si el empleado está en centro_de_costos_prorrateo
        if any(cc[0] == employee_data['ME01010.000-ISGDDMFNN'].iloc[0] for cc in centro_de_costos_prorrateo): #ME01010.000-ISGDDMFNN
            for centro_de_costos in centro_de_costos_prorrateo:
                percentage_cc = centro_de_costos[1]
                cost_center = centro_de_costos[2]  # Usar el CC correspondiente

                # Solo se procesan los paycodes que están en el DataFrame de base_amounts
                amounts = calcular_amount(value, percentage_cc, df_catalogo_filtrado, paycode, employee_id, df_catalogo3, period, entity, pay_group, work_agreement, employee_name, cost_center, level2, level3, level4)

                # Añadir registros a la lista de análisis
                for _, amount_row in amounts.iterrows():
                    new_row = crear_fila(
                        period, entity, pay_group, work_agreement, employee_id, employee_name,
                        amount_row['Client Code'], amount_row['Campo'], cost_center,  # Cost Center real
                        percentage_cc,
                        level2, level3, level4, amount_row['Amount'], amount_row['DEBIT account'], amount_row['CREDIT account'],
                        amount_row.get('Item Text', ''),  # Item Text obtenido del cálculo o valor por defecto
                        amount_row['GLFile'])
                    analysis.append(new_row)

        else:  # Si no está en centro_de_costos_prorrateo, usar centro_de_costos_solo
            for centro_de_costos in centro_de_costos_solo:
                percentage_cc = centro_de_costos[1]
                cost_center = centro_de_costos[2]  # Usar el CC correspondiente

                # Crear una clave única para verificar duplicados
                key = (employee_data['ME01010.000-ISGDDMFNN'], paycode, percentage_cc) #ME01010.000-ISGDDMFNN

                # Solo proceder si no hemos procesado esta combinación antes
                if key not in processed_keys:
                    processed_keys.add(key)  # Marcar como procesado

                    # Solo se procesan los paycodes que están en el DataFrame de base_amounts
                    amounts = calcular_amount(value, percentage_cc, df_catalogo_filtrado, paycode, employee_id, df_catalogo3, period, entity, pay_group, work_agreement, employee_name, cost_center, level2, level3, level4)

                    # Añadir registros a la lista de análisis
                    for _, amount_row in amounts.iterrows():
                        new_row = crear_fila(
                            period, entity, pay_group, work_agreement, employee_id, employee_name,
                            amount_row['Client Code'], amount_row['Campo'], cost_center,  # Cost Center real
                            percentage_cc,
                            level2, level3, level4, amount_row['Amount'], amount_row['DEBIT account'], amount_row['CREDIT account'],
                            amount_row.get('Item Text', ''),  # Item Text obtenido del cálculo o valor por defecto
                            amount_row['GLFile']
                        )
                        analysis.append(new_row)
    
    return analysis

# Procesar todos los empleados
analysis = []

#Función para filtrar el catalogo
def obtener_catalogo_filtrado(df_catalogo_1, df_prorrateo, period_ptype):
    company_list = obtener_entidad(df_prorrateo)
    
    # Convertir la columna 'Company' a numérico (si es posible)
    df_catalogo_1['Company'] = pd.to_numeric(df_catalogo_1['Company'], errors='coerce')
    
    # Crear la columna 'Pivot look' en todo el DataFrame
    df_catalogo_1['Pivot look'] = (
        df_catalogo_1['Accounting type'].astype(str) + 
        df_catalogo_1['Petal Code'].astype(str) + 
        df_catalogo_1['Company'].astype(str) +
        df_catalogo_1['Process Type'].astype(str) + 
        df_catalogo_1['GLFile'].astype(str) + 
        df_catalogo_1['Client Code'].astype(str)
    )
    
    # Filtrar por Company
    if 'ALL' in company_list:
        company_filter = df_catalogo_1['Company'].isin(company_list) | df_catalogo_1['Company'].isna()
    else:
        company_filter = df_catalogo_1['Company'].isin(company_list)
    
    # Filtrar por Process Type
    if 'ALL' in period_ptype:
        process_type_filter = df_catalogo_1['Process Type'].isin(period_ptype)
    else:
        process_type_filter = df_catalogo_1['Process Type'].isin(period_ptype)
    
    # Aplicar los filtros
    df_filtrado = df_catalogo_1[process_type_filter & company_filter]
    
    # Retornar los registros de df_catalogo_1 que tengan un 'Pivot look' presente en df_filtrado
    return df_catalogo_1[df_catalogo_1['Pivot look'].isin(df_filtrado['Pivot look'].unique())]


# Llamar a la función con los DataFrames y parámetros necesarios
df_catalogo_filtrado = obtener_catalogo_filtrado(df_catalogo1, df_input, period_data)

analysis = []
for employee in df_input['ME01010.000-ISGDDMFNN'].unique():# ME01010.000-ISGDDMFNN
    employee_data = df_input[df_input['ME01010.000-ISGDDMFNN'] == employee].copy() #ME01010.000-ISGDDMFNN
    analysis.extend(procesar_empleado(employee_data, df_catalogo_filtrado, df_catalogo3))

df_analysis = pd.DataFrame(analysis)
df_analysis = df_analysis[df_analysis['Amount'] != 0]
df_detailed = pd.concat([df_detailed, df_analysis], ignore_index=True)

df_detailed_copy = df_detailed.copy()

grouped_columns = [
    "Payroll Period", "Entity", "Pay group", "Client ID", "Client Description", "Cost Center", "Level2 (Optional)",
    "Level3 (Optional)", "Level4 (Optional)", "Amount", "Debt Account", "Acred Account", "Item Text", "GL File", 
    "Observations"
]
df_grouped = pd.DataFrame(columns=grouped_columns)

# Merge accounts and clean empty cells
df_detailed_copy["Debt Account"] = df_detailed_copy["Debt Account"].replace("", np.nan)
df_detailed_copy["Acred Account"] = df_detailed_copy["Acred Account"].replace("", np.nan)
df_detailed_copy["Account"] = df_detailed_copy["Debt Account"].combine_first(df_detailed_copy["Acred Account"])
df_detailed_copy["Account Type"] = np.where(
    df_detailed_copy["Debt Account"].notna(), "D+",
    np.where(df_detailed_copy["Acred Account"].notna(), "C+", None)
)

# Group and aggregate the detailed data
df_grouped_final = df_detailed_copy.groupby(
    ['Payroll Period', 'Entity', 'Pay group', "Client ID", "Client Description", 
     "Cost Center", "Account", "Item Text", "GL File"],
    dropna=False,
    as_index=False
).agg({
    "Level2 (Optional)": 'sum',
    "Level3 (Optional)": 'sum',
    "Level4 (Optional)": 'sum',
    "Amount": 'sum',
    "Account Type": lambda x: x.dropna().iloc[0] if not x.dropna().empty else None
})

df_grouped_final["Debt Account"] = np.where(df_grouped_final["Account Type"] == "D+", 
                                            df_grouped_final["Account"], '')
df_grouped_final["Acred Account"] = np.where(df_grouped_final["Account Type"] == "C+", 
                                             df_grouped_final["Account"], '')

df_grouped_final = df_grouped_final[['Payroll Period','Entity','Pay group','Client ID','Client Description',
                                     'Cost Center','Level2 (Optional)','Level3 (Optional)','Level4 (Optional)',
                                     'Amount','Debt Account','Acred Account','Item Text','GL File']]

df_grouped = pd.concat([df_grouped, df_grouped_final])


def data_format(df_detailed, df_grouped, folder_location, file_location, new_file_location, country):
    """
    Transfers data from the input and catalogs sheets of the source Excel file to the output Excel file.
    """
    wb = openpyxl.load_workbook(file_location, keep_vba=True)
    sheet_input_in = wb['Input']
    sheet_catalog = wb['Catalog']

    ob = openpyxl.load_workbook(new_file_location, keep_vba=True)
    sheet_input_out = ob['Input']
    sheet_catalog_out = ob['Catalog']
    sheet_detailed = ob['Detailed']
    sheet_grouped = ob['Grouped']

    # File naming format with entity and period
    id_entidad = int(sheet_input_in['B2'].value)
    dat = datetime.now().strftime('%Y%m%d')
    periodo = sheet_input_in['A2'].value.replace(' ', '').replace('/', '').replace('-', '')

    ruta = os.path.join(os.getcwd(), folder_location)
    carpeta = os.path.join(ruta, f"GL_{country}_{id_entidad}_{periodo}_{dat}")

    if not os.path.exists(carpeta):
        os.makedirs(carpeta)

    s_filename = os.path.join(carpeta, f"GL_{country}_{id_entidad}_{periodo}_{dat}.xlsm")

    # Paste df_detailed data into 'Detailed' sheet
    for r_idx, row in df_detailed.iterrows():
        for c_idx, value in enumerate(row):
            sheet_detailed.cell(row=r_idx + 6, column=c_idx + 2, value=value)

    # Paste df_grouped data into 'Grouped' sheet
    for r_idx, row in df_grouped.iterrows():
        for c_idx, value in enumerate(row):
            sheet_grouped.cell(row=r_idx + 6, column=c_idx + 2, value=value)

    # Copy column names and data from the 'Catalogos' sheet
    column_names_catalog = list(sheet_catalog.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    for c_idx, value in enumerate(column_names_catalog):
        sheet_catalog_out.cell(row=3, column=c_idx + 1, value=value)
    for r_idx, row in enumerate(sheet_catalog.iter_rows(min_row=4, values_only=True)):
        for c_idx, value in enumerate(row):
            sheet_catalog_out.cell(row=r_idx + 4, column=c_idx + 1, value=value)

    # Copy column names and data from the 'Input' sheet
    column_names_input = list(sheet_input_in.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    for c_idx, value in enumerate(column_names_input):
        sheet_input_out.cell(row=3, column=c_idx + 1, value=value)
    for r_idx, row in enumerate(sheet_input_in.iter_rows(min_row=4, values_only=True)):
        for c_idx, value in enumerate(row):
            sheet_input_out.cell(row=r_idx + 4, column=c_idx + 1, value=value)

    try:
        ob.save(s_filename)
    except Exception as e:
        print(f"Automatic save failed. Error: {e}")
        print(f"Please save the output file manually as: GL_{country}_{id_entidad}_{periodo}_{dat}.xlsm")

# Example usage: 
# Note: Adjust 'folder_location' and 'country' as needed.
data_format(df_detailed, df_grouped, folder_location="output_folder", 
            file_location=FILES_STORAGE, new_file_location=OUTPUT_STORAGE, country='MX')