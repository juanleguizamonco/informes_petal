import openpyxl
import pandas as pd
import datetime
import os

def gl_data_format(
        df1: pd.DataFrame, 
        df2: pd.DataFrame = None, 
        df3: pd.DataFrame = None, 
        storage_location: str = "", 
        file_location: str = "", 
        new_file_location: str = "", 
        country: str = "", 
        entity: str = "", 
        period: str = "", 
        file_type: str = "", 
        input_sheet: str = "", 
        catalog_sheet: str = "", 
        other_sheet1: str = None, 
        other_sheet2: str = None, 
        other_sheet3: str = None
        ):
    """
    Transfers data from the input and catalogs sheets of the source Excel file to the output Excel file.
    Only df1 and input_sheet are required. All other parameters are optional.
    """
    
    # Verificar que df1 (primer DataFrame) esté presente
    if df1 is None or df1.empty:
        raise ValueError("df1 is required and cannot be empty")
    
    # Verificar que los parámetros básicos estén presentes si se van a usar
    if file_location and new_file_location:
        wb = openpyxl.load_workbook(file_location, keep_vba=True)
        ob = openpyxl.load_workbook(new_file_location, keep_vba=True)
        
        # Procesar hoja de input (obligatoria si se especifica)
        if input_sheet:
            try:
                sheet_input_in = wb[input_sheet]
                sheet_input_out = ob[input_sheet]
                
                # Copy column names and data from the 'Input' sheet
                column_names_input = list(sheet_input_in.iter_rows(min_row=3, max_row=3, values_only=True))[0]
                for c_idx, value in enumerate(column_names_input):
                    sheet_input_out.cell(row=3, column=c_idx + 1, value=value)
                for r_idx, row in enumerate(sheet_input_in.iter_rows(min_row=4, values_only=True)):
                    for c_idx, value in enumerate(row):
                        sheet_input_out.cell(row=r_idx + 4, column=c_idx + 1, value=value)
            except KeyError:
                print(f"Warning: Sheet '{input_sheet}' not found in workbooks")
        
        # Procesar hoja de catálogo (opcional)
        if catalog_sheet:
            try:
                sheet_catalog = wb[catalog_sheet]
                sheet_catalog_out = ob[catalog_sheet]
                
                # Copy column names and data from the 'Catalogos' sheet
                column_names_catalog = list(sheet_catalog.iter_rows(min_row=3, max_row=3, values_only=True))[0]
                for c_idx, value in enumerate(column_names_catalog):
                    sheet_catalog_out.cell(row=3, column=c_idx + 1, value=value)
                for r_idx, row in enumerate(sheet_catalog.iter_rows(min_row=4, values_only=True)):
                    for c_idx, value in enumerate(row):
                        sheet_catalog_out.cell(row=r_idx + 4, column=c_idx + 1, value=value)
            except KeyError:
                print(f"Warning: Sheet '{catalog_sheet}' not found in workbooks")
        
        # Procesar primera hoja adicional (opcional)
        if other_sheet1 and df2 is not None and not df2.empty:
            try:
                sheet_1 = ob[other_sheet1]
                # Paste df2 data into first additional sheet
                for r_idx, row in df2.iterrows():
                    for c_idx, value in enumerate(row):
                        sheet_1.cell(row=r_idx + 6, column=c_idx + 2, value=value)
            except KeyError:
                print(f"Warning: Sheet '{other_sheet1}' not found in output workbook")
        
        # Procesar segunda hoja adicional (opcional)
        if other_sheet2 and df3 is not None and not df3.empty:
            try:
                sheet_2 = ob[other_sheet2]
                # Paste df3 data into second additional sheet
                for r_idx, row in df3.iterrows():
                    for c_idx, value in enumerate(row):
                        sheet_2.cell(row=r_idx + 6, column=c_idx + 2, value=value)
            except KeyError:
                print(f"Warning: Sheet '{other_sheet2}' not found in output workbook")
        
        # Generar nombre del archivo y guardarlo (opcional)
        if storage_location and entity and period and country and file_type:
            try:
                # File naming format with entity and period
                id_entidad = df1[entity].iloc[0]
                dat = datetime.datetime.now().strftime('%Y%m%d')
                periodo = df1[period].iloc[0]

                ruta = os.path.join(os.getcwd(), storage_location)
                carpeta = os.path.join(ruta, f"{file_type}_{country}_{id_entidad}_{periodo}_{dat}")

                if not os.path.exists(carpeta):
                    os.makedirs(carpeta)

                s_filename = os.path.join(carpeta, f"{file_type}_{country}_{id_entidad}_{periodo}_{dat}.xlsm")
                
                ob.save(s_filename)
                print(f"File saved successfully as: {s_filename}")
                
            except Exception as e:
                print(f"Automatic save failed. Error: {e}")
                if entity and period and country and file_type:
                    try:
                        id_entidad = df1[entity].iloc[0]
                        periodo = df1[period].iloc[0]
                        dat = datetime.datetime.now().strftime('%Y%m%d')
                        print(f"Please save the output file manually as: {file_type}_{country}_{id_entidad}_{periodo}_{dat}.xlsm")
                    except:
                        print("Please save the output file manually")
        else:
            print("Warning: File will not be saved automatically. Missing required parameters for file naming.")
    
    else:
        print("Warning: No file processing performed. file_location and new_file_location are required for Excel operations.")
    
    return True


# Ejemplo de uso con solo el primer DataFrame (mínimo requerido):
def example_minimal_usage():
    """
    Ejemplo de uso mínimo con solo df1
    """
    import pandas as pd
    
    # Solo el primer DataFrame es obligatorio
    df1 = pd.DataFrame({'entity': ['ENT001'], 'period': ['2024Q1']})
    
    result = gl_data_format(df1=df1)
    return result


# Ejemplo de uso completo:
def example_full_usage():
    """
    Ejemplo de uso completo con todos los parámetros
    """
    import pandas as pd
    
    df1 = pd.DataFrame({'entity': ['ENT001'], 'period': ['2024Q1']})
    df2 = pd.DataFrame({'detailed_data': [1, 2, 3]})
    df3 = pd.DataFrame({'grouped_data': [10, 20, 30]})
    
    result = gl_data_format(
        df1=df1,
        df2=df2,
        df3=df3,
        storage_location="output",
        file_location="source.xlsm",
        new_file_location="template.xlsm",
        country="MX",
        entity="entity",
        period="period",
        file_type="GL",
        input_sheet="Input",
        catalog_sheet="Catalogos",
        other_sheet1="Detailed",
        other_sheet2="Grouped"
    )
    return result